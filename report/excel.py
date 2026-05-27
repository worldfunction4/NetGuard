"""openpyxl Excel 巡检报告模块"""
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import THRESHOLDS


# ── 样式常量 ──────────────────────────────────────────────────────────────────

_HEAD_FILL  = PatternFill("solid", fgColor="1A3A5C")
_WARN_FILL  = PatternFill("solid", fgColor="FFF3CD")
_ERR_FILL   = PatternFill("solid", fgColor="F8D7DA")
_EVEN_FILL  = PatternFill("solid", fgColor="F8FAFC")
_HEAD_FONT  = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=10)
_BODY_FONT  = Font(name="Microsoft YaHei", size=10)
_WARN_FONT  = Font(name="Microsoft YaHei", size=10, bold=True, color="856404")
_ERR_FONT   = Font(name="Microsoft YaHei", size=10, bold=True, color="721C24")
_CENTER     = Alignment(horizontal="center", vertical="center")
_LEFT       = Alignment(horizontal="left", vertical="center")
_THIN       = Side(style="thin", color="D0D7E2")
_BORDER     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def generate_excel_report(metrics: list, output_path: str) -> str:
    """
    将巡检指标列表写入 Excel 文件。

    Args:
        metrics:     巡检结果列表，每项为一台设备的指标字典
                     示例：[{"name": "SW-01", "ip": "1.1.1.1", "cpu_percent": 45, ...}]
        output_path: 输出 .xlsx 文件的完整路径（自动创建父目录）

    Returns:
        输出文件的路径字符串
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "巡检报告"

    # ── 标题行 ─────────────────────────────────────────────────────────────
    title_row = [
        "设备名", "IP 地址", "位置", "角色", "设备类型",
        "CPU %", "内存 %", "接口在线", "接口离线", "状态", "告警详情", "采集时间",
    ]
    ws.append(title_row)
    for col_idx, _ in enumerate(title_row, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEAD_FONT
        cell.fill = _HEAD_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER

    # ── 数据行 ─────────────────────────────────────────────────────────────
    for row_idx, dev in enumerate(metrics, start=2):
        alerts = dev.get("alerts") or []
        status = dev.get("status", "")
        cpu    = dev.get("cpu_percent")
        mem    = dev.get("memory_percent")

        row_data = [
            _safe_cell(dev.get("name", "")),
            _safe_cell(dev.get("ip", "")),
            _safe_cell(dev.get("location", "")),
            _safe_cell(dev.get("role", "")),
            _safe_cell(dev.get("device_type", "")),
            _fmt_pct(cpu),
            _fmt_pct(mem),
            dev.get("interfaces_up", "—"),
            dev.get("interfaces_down", "—"),
            _status_label(status, bool(alerts)),
            _safe_cell("; ".join(a["text"] if isinstance(a, dict) else a for a in alerts) if alerts else ""),
            _safe_cell(dev.get("collected_at", "")),
        ]
        ws.append(row_data)

        # 决定行底色
        is_error = status in ("unreachable", "error")
        is_warn  = bool(alerts)
        fill = _ERR_FILL if is_error else (_WARN_FILL if is_warn else (_EVEN_FILL if row_idx % 2 == 0 else None))

        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _BORDER
            cell.alignment = _CENTER if col_idx not in (1, 3, 4, 11) else _LEFT
            if fill:
                cell.fill = fill

        # CPU / 内存超阈值红字（阈值从 config.THRESHOLDS 读取）
        for col_idx, val, metric_key in [(6, cpu, "cpu_percent"), (7, mem, "memory_percent")]:
            threshold = THRESHOLDS.get(metric_key, 80)
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(val, (int, float)) and val >= threshold:
                cell.font = _WARN_FONT
            else:
                cell.font = _BODY_FONT

    # ── 列宽自适应 ─────────────────────────────────────────────────────────
    col_widths = [16, 16, 14, 10, 18, 8, 8, 9, 9, 10, 40, 18]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 20

    # ── 冻结首行 ───────────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── 汇总 Sheet ─────────────────────────────────────────────────────────
    _write_summary_sheet(wb, metrics)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return str(out)


# ── 汇总 Sheet ────────────────────────────────────────────────────────────────

def _write_summary_sheet(wb: openpyxl.Workbook, metrics: list):
    ws = wb.create_sheet("汇总")
    ok      = sum(1 for d in metrics if d.get("status") == "ok" and not d.get("alerts"))
    warned  = sum(1 for d in metrics if d.get("alerts"))
    err     = sum(1 for d in metrics if d.get("status") in ("unreachable", "error"))

    rows = [
        ["NetGuard 巡检汇总"],
        ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["设备总数", len(metrics)],
        ["正常", ok],
        ["告警", warned],
        ["异常 / 不可达", err],
    ]
    for r in rows:
        ws.append(r)

    ws["A1"].font = Font(name="Microsoft YaHei", bold=True, size=14)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22


# ── 辅助函数 ──────────────────────────────────────────────────────────────────

def _fmt_pct(val) -> str:
    return f"{val}%" if val is not None else "—"


def _status_label(status: str, has_alert: bool) -> str:
    if has_alert:
        return "告警"
    return {"ok": "正常", "unreachable": "不可达", "error": "错误"}.get(status, status)


def _safe_cell(value) -> str:
    """防止 Excel 公式注入（Formula/CSV Injection）。

    以 = + - @ TAB 回车 开头的字符串会被 Excel 解释为公式。
    在前面加英文单引号，Excel 会将整个单元格视为纯文本字符串。
    """
    if not isinstance(value, str):
        return value  # 数字 / None 直接返回，不受影响
    if value and value[0] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + value
    return value
